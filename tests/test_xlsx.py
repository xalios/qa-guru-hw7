from openpyxl import load_workbook
from os.path import join, abspath, dirname


xlsx_path = abspath(
    join(dirname(__file__), './resources/file_example_XLSX_50.xlsx')
)


def test_xlsx():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь

    workbook = load_workbook(xlsx_path)
    sheet = workbook.active

    assert len(sheet.tables) == 0
    assert sheet.cell(row=3, column=2).value == 'Mara'
